import calendar
from datetime import date, timedelta
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import (
    login_required,
    permission_required,
)
from datetime import date, datetime, timedelta
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.core.exceptions import (
    PermissionDenied,
    ValidationError,
)
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from django.core.exceptions import ValidationError

import random
from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from django.contrib.auth import login as django_login
from django.core.mail import send_mail
from django.conf import settings
from .models import EmailOTP
from django.db import transaction

from .forms import (
    AttendanceExceptionForm,
    AttendanceExceptionReviewForm,
    AttendanceUpdateForm,
    EmployeeForm,
    LeaveBalanceForm,
    LeaveRequestForm,
    LeaveRequestReviewForm,
    ProfileFormAdmin,
    SalaryForm,
    ResignationForm,
    ResignationReviewForm,
    ClearanceApprovalDecisionForm,
    TerminationForm,
    HolidayForm,
    HolidayBulkUploadForm,WorkScheduleForm,
    OptionalHolidayRequestForm,


)
from django.db.models import Count, Q
from .services import (
    create_clearance_request,
    submit_resignation,
    review_resignation,
    decide_clearance_approval,
    finalize_employee_exit,
    initiate_termination,
)
import calendar
from datetime import date

from .models import (
    Attendance,
    AttendanceException,
    LATE_CUTOFF_TIME,
    LeaveBalance,
    LeaveRequest,
    LeaveTypeMaster,
    Salary,
    empProfile,
    ClearanceRequest,
    EmployeeExitRequest,
    ClearanceApproval,
    Holiday,
    WorkSchedule,
    OptionalHolidayRequest,
)
import csv
import io
from appEmp.whatsapp_router import find_employee_by_phone

from openpyxl import load_workbook
from datetime import datetime
from .services import create_clearance_request,finalize_employee_exit
from .models import OptionalHolidayRequest
from .models import WhatsAppTicket
import json
from .whatsapp_utils import send_whatsapp_message
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse
from appEmp.whatsapp_router import (
    classify_message,
    find_employee_by_phone,
    get_leave_balance_reply,
    get_attendance_reply,
    get_payroll_reply,
)

from django.views.decorators.cache import never_cache

def user_has_permission(user, permission_codename):
    return (
        user.is_superuser
        or user.has_perm(f"appEmp.{permission_codename}")
    )


def user_has_hr_access(user):
    return user_has_permission(user, "view_hr_dashboard")


def can_review_leave(user):
    return user_has_permission(user, "review_leave")


def can_review_attendance_exception(user):
    return user_has_permission(user, "review_attendance_exception")


@login_required
def empList_view(request):
    if not user_has_permission(request.user, "manage_employees"):
        messages.error(request, "You don't have permission to view employees.")
        return redirect("dashboard")

    objects = empProfile.objects.values(
        "id",
        "user__first_name",
        "user__last_name",
        "phone_number",
        "uuid",
    )

    return render(
        request,
        "appEmp/empList.html",
        {"objects": objects},
    )


@login_required
def profileDetail_view(request, uuid):
    logged_in_profile = get_object_or_404(
        empProfile,
        user=request.user
    )

    employee = get_object_or_404(
        empProfile.objects.select_related(
            "user",
            "manager__user",
            "designation",
            "shift",
        ),
        uuid=uuid,
    )

    # is_admin = user_has_permission(request.user, "manage_employees")
    is_admin = user_has_permission(request.user, "manage_employees")
    is_direct_manager = employee.manager_id == logged_in_profile.id

    if not is_admin and not is_direct_manager:
        messages.error(
            request,
            "You don't have permission to edit this employee profile."
        )
        return redirect("dashboard")

    if request.method == "POST":
        form = ProfileFormAdmin(
            request.POST,
            instance=employee
        )

        if form.is_valid():
            form.save()

            messages.success(
                request,
                "Profile updated successfully."
            )

            return redirect(
                "editProfile",
                uuid=employee.uuid
            )
    else:
        form = ProfileFormAdmin(
            instance=employee
        )

    target_salary = employee.salary_records.filter(
        is_active=True
    ).first()

    return render(
        request,
        "appEmp/profile.html",
        {
            "employee": employee,
            "profile": employee,
            "form": form,
            "salary": target_salary,
            # "can_edit_salary": is_admin,
            "can_edit_salary": user_has_permission(request.user,"manage_salary"),
            "can_edit_profile": True,
            "salary_form": SalaryForm(
                instance=target_salary
            ),
            "target_uuid": employee.uuid,
            "current_year": timezone.now().year,
            "current_month": timezone.now().month,
        },
    )


@login_required
def empProfile_view(request):
    profile, created = empProfile.objects.get_or_create(user=request.user)
    if request.method == "POST":
        form = EmployeeForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, "Profile updated successfully.")
            return redirect("myProfile")
    else:
        form = EmployeeForm(instance=profile)

    self_salary = profile.salary_records.filter(is_active=True).first()

    return render(
        request,
        "appEmp/profile.html",
        {
            "form": form,
            "profile": profile,
            "created": created,
            "salary": self_salary,
            "can_edit_salary": False,
            "target_uuid": profile.uuid,
            "current_year": timezone.now().year,
            "current_month": timezone.now().month,
        },
    )

@never_cache
@login_required
def dashboard_view(request):
    profile = get_object_or_404(
        empProfile.objects.select_related(
            "user",
            "manager__user",
            "shift",
            "work_schedule",
            "designation",
        ),
        user=request.user,
    )

    today = timezone.localdate()
    now = timezone.localtime()

    is_admin_view = user_has_hr_access(
        request.user
    )

    # =========================================================
    # EXIT / CLEARANCE — COMMON COUNTS
    # =========================================================

    active_resignation_count = (
        EmployeeExitRequest.objects.filter(
            employee=profile,
            exit_type="RESIGNATION",
            status__in=[
                "SUBMITTED",
                "CLEARANCE_IN_PROGRESS",
                "READY_FOR_FINALIZATION",
            ],
        ).count()
    )

    pending_manager_noc_count = (
        ClearanceApproval.objects.filter(
            approval_type="MANAGER_NOC",
            assigned_to=request.user,
            status="PENDING",
            clearance_request__status="PENDING",
        ).count()
    )

    pending_finance_noc_count = 0

    if (
        request.user.is_superuser
        or request.user.has_perm(
            "appEmp.review_finance_noc"
        )
    ):
        pending_finance_noc_count = (
            ClearanceApproval.objects.filter(
                approval_type="FINANCE_NOC",
                status="PENDING",
                clearance_request__status="PENDING",
            ).count()
        )

    pending_resignations_count = 0
    ready_for_finalization_count = 0
    rejected_exit_requests_count = 0
    total_exit_requests_count = 0

    if (
        request.user.is_superuser
        or request.user.has_perm(
            "appEmp.view_all_clearance_requests"
        )
    ):
        pending_resignations_count = (
            EmployeeExitRequest.objects.filter(
                exit_type="RESIGNATION",
                status="SUBMITTED",
            ).count()
        )

        rejected_exit_requests_count = (
            EmployeeExitRequest.objects.filter(
                status__in=[
                    "HR_REJECTED",
                    "CLEARANCE_REJECTED",
                ],
            ).count()
        )

        total_exit_requests_count = (
            EmployeeExitRequest.objects.count()
        )

    if (
        request.user.is_superuser
        or request.user.has_perm(
            "appEmp.finalize_employee_exit"
        )
    ):
        ready_for_finalization_count = (
            EmployeeExitRequest.objects.filter(
                status="READY_FOR_FINALIZATION",
            ).count()
        )

    exiting_employees_count = (
        EmployeeExitRequest.objects.filter(
            status__in=[
                "SUBMITTED",
                "CLEARANCE_IN_PROGRESS",
                "READY_FOR_FINALIZATION",
            ]
        )
        .values(
            "employee_id"
        )
        .distinct()
        .count()
    )

    exit_clearance_context = {
        "active_resignation_count":
            active_resignation_count,

        "pending_manager_noc_count":
            pending_manager_noc_count,

        "pending_finance_noc_count":
            pending_finance_noc_count,

        "pending_resignations_count":
            pending_resignations_count,

        "ready_for_finalization_count":
            ready_for_finalization_count,

        "rejected_exit_requests_count":
            rejected_exit_requests_count,

        "total_exit_requests_count":
            total_exit_requests_count,

        "exiting_employees_count":
            exiting_employees_count,
    }

    # =========================================================
    # HR / ADMIN DASHBOARD
    # =========================================================

    if is_admin_view:

        # -----------------------------------------------------
        # WORKFORCE
        # -----------------------------------------------------

        total_employees = (
            empProfile.objects.count()
        )

        active_employees = (
            empProfile.objects.filter(
                is_active=True
            ).count()
        )

        recent_employees = (
            empProfile.objects
            .filter(
                is_active=True
            )
            .select_related(
                "user",
                "designation",
            )
            .order_by(
                "-user__date_joined"
            )[:5]
        )

        # -----------------------------------------------------
        # VERIFICATION
        # -----------------------------------------------------

        verification_stats = {
            "email": (
                empProfile.objects.filter(
                    is_email_verified=True
                ).count()
            ),

            "phone": (
                empProfile.objects.filter(
                    is_phone_verified=True
                ).count()
            ),

            "address": (
                empProfile.objects.filter(
                    is_address_verified=True
                ).count()
            ),

            "aadhar": (
                empProfile.objects.filter(
                    is_aadhar_verified=True
                ).count()
            ),

            "background_check": (
                empProfile.objects.filter(
                    is_background_check_completed=True
                ).count()
            ),
        }

        # -----------------------------------------------------
        # TODAY ATTENDANCE
        # -----------------------------------------------------

        today_attendance = (
            Attendance.objects
            .filter(
                date=today
            )
            .select_related(
                "employee__user"
            )
        )

        present_today = (
            today_attendance.filter(
                status__in=[
                    "PRESENT",
                    "LATE",
                    "HALF_DAY",
                ]
            ).count()
        )

        absent_today = (
            today_attendance.filter(
                status="ABSENT"
            ).count()
        )
        # -----------------------------------------------------
        # ATTENDANCE RECORD TABLE PERIOD
        # -----------------------------------------------------

        attendance_period = request.GET.get(
            "attendance_period",
            "today",
        )

        allowed_attendance_periods = {
            "today",
            "7",
            "30",
        }

        if attendance_period not in allowed_attendance_periods:
            attendance_period = "today"


        if attendance_period == "7":

            attendance_start_date = (
                today - timedelta(days=6)
            )

            attendance_period_label = (
                "Last 7 Days"
            )

        elif attendance_period == "30":

            attendance_start_date = (
                today - timedelta(days=29)
            )

            attendance_period_label = (
                "Last 30 Days"
            )

        else:

            attendance_start_date = today

            attendance_period_label = (
                "Today"
            )


        attendance_records = (
            Attendance.objects
            .filter(
                date__gte=attendance_start_date,
                date__lte=today,
            )
            .select_related(
                "employee__user"
            )
            .order_by(
                "-date",
                "employee__user__first_name",
            )
        )

        # Approved LeaveRequest is the real leave source.
        approved_leave_today = (
            LeaveRequest.objects
            .filter(
                status="APPROVED",
                start_date__lte=today,
                end_date__gte=today,
            )
            .select_related(
                "employee__user",
                "leave_type",
            )
        )

        leave_employee_ids = set(
            approved_leave_today.values_list(
                "employee_id",
                flat=True,
            )
        )

        attendance_employee_ids = set(
            today_attendance.values_list(
                "employee_id",
                flat=True,
            )
        )

        employees_on_leave_today = (
            len(
                leave_employee_ids
            )
        )

        accounted_employee_ids = (
            attendance_employee_ids
            | leave_employee_ids
        )

        not_marked_today = max(
            active_employees
            - len(accounted_employee_ids),
            0,
        )

        attendance_denominator = max(
            active_employees
            - employees_on_leave_today,
            1,
        )

        attendance_percentage = min(
            round(
                (
                    present_today
                    / attendance_denominator
                )
                * 100
            ),
            100,
        )

        attendance_summary = {
            "present": present_today,
            "absent": absent_today,
            "leave": employees_on_leave_today,
            "not_marked": not_marked_today,
        }

        # -----------------------------------------------------
        # 7-DAY ATTENDANCE TREND
        # -----------------------------------------------------

        trend_start = (
            today
            - timedelta(days=6)
        )

        trend_query = (
            Attendance.objects
            .filter(
                date__gte=trend_start,
                date__lte=today,
            )
            .values(
                "date"
            )
            .annotate(
                present_count=Count(
                    "id",
                    filter=Q(
                        status__in=[
                            "PRESENT",
                            "LATE",
                            "HALF_DAY",
                        ]
                    ),
                ),

                absent_count=Count(
                    "id",
                    filter=Q(
                        status="ABSENT"
                    ),
                ),
            )
            .order_by(
                "date"
            )
        )

        trend_map = {
            item["date"]: item
            for item in trend_query
        }

        attendance_trend = []

        for offset in range(7):

            trend_date = (
                trend_start
                + timedelta(
                    days=offset
                )
            )

            values = trend_map.get(
                trend_date,
                {
                    "present_count": 0,
                    "absent_count": 0,
                },
            )

            present_count = (
                values["present_count"]
            )

            absent_count = (
                values["absent_count"]
            )

            if active_employees:
                present_percent = round(
                    (
                        present_count
                        / active_employees
                    )
                    * 100
                )
            else:
                present_percent = 0

            attendance_trend.append({
                "date": trend_date,

                "label": (
                    trend_date.strftime(
                        "%a"
                    )
                ),

                "present": present_count,

                "absent": absent_count,

                "present_percent":
                    present_percent,
            })

        # -----------------------------------------------------
        # PENDING HR ACTIONS
        # -----------------------------------------------------

        pending_leave_requests_count = (
            LeaveRequest.objects.filter(
                status="PENDING"
            ).count()
        )

        pending_attendance_exceptions_count = (
            AttendanceException.objects.filter(
                status="PENDING"
            ).count()
        )

        pending_optional_holiday_requests_count = (
            OptionalHolidayRequest.objects.filter(
                status="PENDING"
            ).count()
        )

        current_year = (
            today.year
        )

        employees_without_leave_balance = (
            empProfile.objects
            .filter(
                is_active=True
            )
            .exclude(
                leave_balances__year=current_year
            )
            .distinct()
            .count()
        )

        pending_hr_actions_count = (
            pending_leave_requests_count
            + pending_attendance_exceptions_count
            + pending_optional_holiday_requests_count
            + pending_resignations_count
            + pending_manager_noc_count
            + pending_finance_noc_count
            + ready_for_finalization_count
        )

        # -----------------------------------------------------
        # LEAVE OVERVIEW
        # -----------------------------------------------------

        upcoming_approved_leaves = (
            LeaveRequest.objects
            .filter(
                status="APPROVED",
                end_date__gte=today,
            )
            .select_related(
                "employee__user",
                "leave_type",
            )
            .order_by(
                "start_date"
            )[:5]
        )

        employees_on_leave_today_list = (
            approved_leave_today
            .order_by(
                "employee__user__first_name",
                "employee__user__last_name",
            )[:5]
        )

        # -----------------------------------------------------
        # UPCOMING HOLIDAYS
        # -----------------------------------------------------

        upcoming_holidays = (
            Holiday.objects
            .filter(
                is_active=True,
                date__gte=today,
            )
            .order_by(
                "date"
            )[:5]
        )

        # -----------------------------------------------------
        # CONTEXT
        # -----------------------------------------------------

        context = {
            "is_admin_view": True,

            "total_employees":
                total_employees,

            "active_employees":
                active_employees,

            "recent_employees":
                recent_employees,

            "verification_stats":
                verification_stats,

            "attendance_summary":
                attendance_summary,

            "attendance_percentage":
                attendance_percentage,

            "attendance_trend":
                attendance_trend,

            "today_attendance":
                today_attendance,

            "employees_on_leave_today":
                employees_on_leave_today,

            "employees_on_leave_today_list":
                employees_on_leave_today_list,

            "upcoming_approved_leaves":
                upcoming_approved_leaves,

            "pending_leave_requests_count":
                pending_leave_requests_count,

            "pending_attendance_exceptions_count":
                pending_attendance_exceptions_count,

            "pending_optional_holiday_requests_count":
                pending_optional_holiday_requests_count,

            "employees_without_leave_balance":
                employees_without_leave_balance,

            "pending_hr_actions_count":
                pending_hr_actions_count,

            "upcoming_holidays":
                upcoming_holidays,

            "attendance_records": (
                attendance_records
            ),

            "attendance_period": (
                attendance_period
            ),

            "attendance_period_label": (
                attendance_period_label
            ),
        }

    # =========================================================
    # EMPLOYEE DASHBOARD
    # =========================================================

    else:

        # Night-shift aware attendance date.
        attendance_date = (
            _get_attendance_date(
                profile,
                now,
            )
        )

        day_context = (
            _get_employee_day_context(
                profile,
                attendance_date,
            )
        )

        my_attendance_today = (
            Attendance.objects
            .filter(
                employee=profile,
                date=attendance_date,
            )
            .first()
        )

        recent_attendance = (
            Attendance.objects
            .filter(
                employee=profile
            )
            .order_by(
                "-date"
            )[:7]
        )

        my_pending_leave_requests = (
            LeaveRequest.objects
            .filter(
                employee=profile,
                status="PENDING",
            )
            .count()
        )

        my_upcoming_leaves = (
            LeaveRequest.objects
            .filter(
                employee=profile,
                status="APPROVED",
                end_date__gte=today,
            )
            .select_related(
                "leave_type"
            )
            .order_by(
                "start_date"
            )[:3]
        )

        my_upcoming_optional_holidays = (
            OptionalHolidayRequest.objects
            .filter(
                employee=profile,
                status="APPROVED",
                holiday__is_active=True,
                holiday__date__gte=today,
            )
            .select_related(
                "holiday"
            )
            .order_by(
                "holiday__date"
            )[:3]
        )

        upcoming_holidays = (
            Holiday.objects
            .filter(
                is_active=True,
                holiday_type="MANDATORY",
                date__gte=today,
            )
            .order_by(
                "date"
            )[:3]
        )

        context = {
            "is_admin_view": False,

            "profile": profile,

            "attendance_date":
                attendance_date,

            "my_attendance_today":
                my_attendance_today,

            "recent_attendance":
                recent_attendance,

            "day_context":
                day_context,

            "my_pending_leave_requests":
                my_pending_leave_requests,

            "my_upcoming_leaves":
                my_upcoming_leaves,

            "my_upcoming_optional_holidays":
                my_upcoming_optional_holidays,

            "upcoming_holidays":
                upcoming_holidays,
        }

    # =========================================================
    # COMMON EXIT / CLEARANCE CONTEXT
    # =========================================================

    context.update(
        exit_clearance_context
    )

    return render(
        request,
        "dashboard.html",
        context,
    )



def _get_attendance_date(employee, current_datetime):    
    local_datetime = timezone.localtime(
        current_datetime
    )

    attendance_date = local_datetime.date()

    shift = employee.shift

    if (
        shift
        and shift.is_night_shift
        and local_datetime.time() <= shift.end_time
    ):
        attendance_date = (
            attendance_date
            - timedelta(days=1)
        )

    return attendance_date


def _is_employee_late(employee, check_in_datetime):
    """
    Determine whether an employee is late based on
    assigned shift start time + grace period.

    Supports normal and overnight/night shifts.
    """

    shift = employee.shift

    if not shift:
        return False

    local_check_in = timezone.localtime(
        check_in_datetime
    )

    shift_date = local_check_in.date()

    # For overnight shifts like 22:00–06:00:
    # a 01:00 check-in belongs to the shift
    # that started on the previous calendar day.
    if (
        shift.is_night_shift
        and local_check_in.time() <= shift.end_time
    ):
        shift_date = (
            shift_date
            - timedelta(days=1)
        )

    shift_start_datetime = datetime.combine(
        shift_date,
        shift.start_time,
    )

    shift_start_datetime = timezone.make_aware(
        shift_start_datetime,
        timezone.get_current_timezone(),
    )

    late_cutoff = (
        shift_start_datetime
        + timedelta(
            minutes=shift.grace_period_minutes
        )
    )

    return local_check_in > late_cutoff


# def _get_employee_day_context(employee, target_date):
#     """
#     Resolve the employee's work status for a specific date.

#     Priority:
#     1. Mandatory Holiday
#     2. Weekly Off
#     3. Approved Leave
#     4. Normal Working Day

#     Optional holidays do not automatically count as off.
#     """

#     # --------------------------------------------------
#     # MANDATORY HOLIDAY
#     # --------------------------------------------------

#     holiday = Holiday.objects.filter(
#         date=target_date,
#         holiday_type="MANDATORY",
#         is_active=True,
#     ).first()

#     # --------------------------------------------------
#     # WORK SCHEDULE / WEEKLY OFF
#     # --------------------------------------------------

#     work_schedule = employee.work_schedule

#     if not work_schedule:
#         work_schedule = (
#             WorkSchedule.objects
#             .filter(
#                 is_default=True,
#                 is_active=True,
#             )
#             .first()
#         )

#     is_weekly_off = False

#     if work_schedule:
#         is_weekly_off = not (
#             work_schedule.is_working_day(
#                 target_date
#             )
#         )

#     # --------------------------------------------------
#     # APPROVED LEAVE
#     # --------------------------------------------------

#     approved_leave = (
#         LeaveRequest.objects
#         .filter(
#             employee=employee,
#             status="APPROVED",
#             start_date__lte=target_date,
#             end_date__gte=target_date,
#         )
#         .select_related(
#             "leave_type"
#         )
#         .first()
#     )

#     return {
#         "holiday": holiday,
#         "work_schedule": work_schedule,
#         "is_weekly_off": is_weekly_off,
#         "approved_leave": approved_leave,
#     }

def _get_employee_day_context(employee, target_date):
    """
    Resolve the employee's work status for a specific date.

    Priority:
    1. Mandatory Holiday
    2. Weekly Off
    3. Approved Optional Holiday
    4. Approved Leave
    5. Normal Working Day

    Optional holidays only count as off when the
    employee's request has been approved.
    """

    # ==================================================
    # MANDATORY HOLIDAY
    # ==================================================

    holiday = (
        Holiday.objects
        .filter(
            date=target_date,
            holiday_type="MANDATORY",
            is_active=True,
        )
        .first()
    )

    # ==================================================
    # WORK SCHEDULE
    # ==================================================

    work_schedule = employee.work_schedule

    if not work_schedule:
        work_schedule = (
            WorkSchedule.objects
            .filter(
                is_default=True,
                is_active=True,
            )
            .first()
        )

    is_weekly_off = False

    if work_schedule:
        is_weekly_off = not (
            work_schedule.is_working_day(
                target_date
            )
        )

    # ==================================================
    # APPROVED OPTIONAL HOLIDAY
    # ==================================================

    approved_optional_holiday = (
        OptionalHolidayRequest.objects
        .filter(
            employee=employee,
            holiday__date=target_date,
            holiday__holiday_type="OPTIONAL",
            holiday__is_active=True,
            status="APPROVED",
        )
        .select_related(
            "holiday"
        )
        .first()
    )

    # ==================================================
    # APPROVED NORMAL LEAVE
    # ==================================================

    approved_leave = (
        LeaveRequest.objects
        .filter(
            employee=employee,
            status="APPROVED",
            start_date__lte=target_date,
            end_date__gte=target_date,
        )
        .select_related(
            "leave_type"
        )
        .first()
    )

    return {
        "holiday": holiday,
        "work_schedule": work_schedule,
        "is_weekly_off": is_weekly_off,
        "approved_optional_holiday": (
            approved_optional_holiday
        ),
        "approved_leave": approved_leave,
    }


@login_required
def checkin_checkout_view(request):
    profile = get_object_or_404(
        empProfile.objects.select_related(
            "shift",
            "work_schedule",
        ),
        user=request.user,
    )

    now = timezone.localtime()

    # --------------------------------------------------
    # Resolve attendance date
    # Handles night shifts crossing midnight.
    # --------------------------------------------------
    attendance_date = _get_attendance_date(
        profile,
        now,
    )

    # --------------------------------------------------
    # Resolve holiday / weekly off / leave
    # for the actual attendance date.
    # --------------------------------------------------
    day_context = _get_employee_day_context(
        profile,
        attendance_date,
    )

    # holiday = day_context["holiday"]
    # is_weekly_off = day_context["is_weekly_off"]
    # approved_leave = day_context["approved_leave"]
    holiday = day_context["holiday"]

    is_weekly_off = day_context[
        "is_weekly_off"
    ]

    approved_optional_holiday = day_context[
        "approved_optional_holiday"
    ]

    approved_leave = day_context[
        "approved_leave"
    ]

    # ==================================================
    # MANDATORY HOLIDAY
    # ==================================================

    if holiday:
        messages.info(
            request,
            (
                f"{holiday.name} is a mandatory holiday. "
                "Attendance is not required."
            ),
        )

        return redirect("dashboard")

    # ==================================================
    # WEEKLY OFF
    # ==================================================

    if is_weekly_off:
        messages.info(
            request,
            "This is your weekly off. Attendance is not required.",
        )

        return redirect("dashboard")

    # ==================================================
# APPROVED OPTIONAL HOLIDAY
# ==================================================

    if approved_optional_holiday:
        messages.info(
            request,
            (
                "You have an approved optional holiday: "
                f"{approved_optional_holiday.holiday.name}. "
                "Attendance is not required."
            ),
        )

        return redirect(
            "dashboard"
        )

    # ==================================================
    # APPROVED FULL-DAY LEAVE
    # ==================================================

    if (
        approved_leave
        and approved_leave.duration_type == "FULL_DAY"
    ):
        messages.info(
            request,
            (
                "You have approved "
                f"{approved_leave.leave_type.name} "
                "for this attendance date."
            ),
        )

        return redirect("dashboard")

    # ==================================================
    # GET / CREATE ATTENDANCE
    # ==================================================

    attendance, _ = Attendance.objects.get_or_create(
        employee=profile,
        date=attendance_date,
    )

    # ==================================================
    # CHECK IN
    # ==================================================

    if attendance.check_in is None:
        attendance.check_in = now

        # ----------------------------------------------
        # HALF-DAY APPROVED LEAVE
        # ----------------------------------------------

        if (
            approved_leave
            and approved_leave.duration_type
            in [
                "FIRST_HALF",
                "SECOND_HALF",
            ]
        ):
            attendance.status = "HALF_DAY"

        # ----------------------------------------------
        # NORMAL WORKING DAY
        # ----------------------------------------------

        else:
            attendance.status = (
                "LATE"
                if _is_employee_late(
                    profile,
                    now,
                )
                else "PRESENT"
            )

        attendance.save()

        messages.success(
            request,
            (
                f"Checked in at "
                f"{now.strftime('%I:%M %p')}."
            ),
        )

    # ==================================================
    # CHECK OUT
    # ==================================================

    elif attendance.check_out is None:
        attendance.check_out = now
        attendance.save()

        messages.success(
            request,
            (
                f"Checked out at "
                f"{now.strftime('%I:%M %p')}."
            ),
        )

    # ==================================================
    # ALREADY COMPLETED
    # ==================================================

    else:
        messages.info(
            request,
            "You've already checked in and out for this shift.",
        )

    return redirect("dashboard")


@login_required
def update_attendance_view(request, uuid):
    if not user_has_permission(request.user, "manage_attendance"):
        messages.error(
            request,
            "You don't have permission to do that."
        )
        return redirect("dashboard")

    attendance = get_object_or_404(
        Attendance,
        uuid=uuid,
    )

    if request.method == "POST":
        form = AttendanceUpdateForm(
            request.POST,
            instance=attendance,
        )

        if form.is_valid():
            record = form.save(commit=False)
            record.marked_by = request.user
            record.save()

            messages.success(
                request,
                f"Attendance updated for {attendance.employee}."
            )
            return redirect("dashboard")

    else:
        form = AttendanceUpdateForm(instance=attendance)

    return render(
        request,
        "appEmp/edit_attendance.html",
        {
            "form": form,
            "attendance": attendance,
        },
    )


@login_required
def salary_update_view(request, uuid):
    if not user_has_permission(request.user, "manage_salary"):
        messages.error(
            request,
            "You don't have permission to do this."
        )
        return redirect("dashboard")

    target = get_object_or_404(
        empProfile,
        uuid=uuid,
    )

    if request.method == "POST":
        form = SalaryForm(request.POST)

        if form.is_valid():
            new_salary = form.save(commit=False)
            new_salary.employee = target
            new_salary.is_active = True
            new_salary.created_by = request.user
            new_salary.save()

            messages.success(
                request,
                (
                    "Salary updated for "
                    f"{target.user.get_full_name() or target.user.username}."
                ),
            )
        else:
            messages.error(
                request,
                "Please correct the errors in the salary form."
            )

    return redirect(
        "editProfile",
        uuid=target.uuid,
    )


def _build_salary_slip_pdf(salary, profile, month, year):
    """Renders a salary slip to an in-memory PDF buffer using ReportLab
    (no system dependencies needed, unlike WeasyPrint)."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=20 * mm, bottomMargin=20 * mm,
        leftMargin=20 * mm, rightMargin=20 * mm,
    )
    styles = getSampleStyleSheet()
    elements = []

    NAVY = colors.HexColor('#0F172A')
    TEAL = colors.HexColor('#0F9E96')
    ROSE = colors.HexColor('#9F1239')
    GRID = colors.HexColor('#CBD5E1')

    title_style = ParagraphStyle('SlipTitle', parent=styles['Heading1'], textColor=NAVY, spaceAfter=2)
    elements.append(Paragraph("Salary Slip", title_style))
    elements.append(Paragraph(f"{calendar.month_name[month]} {year}", styles['Normal']))
    elements.append(Spacer(1, 8 * mm))

    emp_info = [
        ['Employee Name', profile.user.get_full_name() or profile.user.username],
        ['Employee ID', str(profile.uuid)[:8].upper()],
        ['Position', profile.position or '—'],
        ['Date of Joining', str(profile.date_hired) if profile.date_hired else '—'],
    ]
    t1 = Table(emp_info, colWidths=[55 * mm, 105 * mm])
    t1.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TEXTCOLOR', (0, 0), (0, -1), TEAL),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
    ]))
    elements.append(t1)
    elements.append(Spacer(1, 8 * mm))

    tds_monthly = salary.calculate_monthly_tds()
    net_pay = salary.net_monthly_pay

    earnings = [
        ['Earnings', 'Amount (INR)'],
        ['Basic Salary', f"{salary.basic_salary:,.2f}"],
        ['HRA', f"{salary.hra:,.2f}"],
        ['Gross Earnings', f"{salary.gross_monthly:,.2f}"],
    ]
    deductions = [
        ['Deductions', 'Amount (INR)'],
        ['Provident Fund (PF)', f"{salary.pf:,.2f}"],
        ['TDS (Tax Deducted at Source)', f"{tds_monthly:,.2f}"],
        ['Total Deductions', f"{(salary.pf + tds_monthly):,.2f}"],
    ]

    t2 = Table(earnings, colWidths=[110 * mm, 50 * mm])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), NAVY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, GRID),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    elements.append(t2)
    elements.append(Spacer(1, 6 * mm))

    t3 = Table(deductions, colWidths=[110 * mm, 50 * mm])
    t3.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), ROSE),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, GRID),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    elements.append(t3)
    elements.append(Spacer(1, 8 * mm))

    net_table = Table([['Net Pay', f"Rs. {net_pay:,.2f}"]], colWidths=[110 * mm, 50 * mm])
    net_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), TEAL),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    elements.append(net_table)
    elements.append(Spacer(1, 10 * mm))

    elements.append(
        Paragraph(
            "This is a system-generated salary slip. "
            "Tax figures shown are an estimate; "
            "refer to Form 16 issued by HR for final tax liability.",
            styles["Italic"],
        )
    )

    doc.build(elements)
    buffer.seek(0)
    return buffer


@login_required
def salary_slip_self_view(request, year, month):
    profile = request.user.empprofile
    salary = profile.salary_records.filter(is_active=True).first()
    if not salary:
        messages.error(request, "No salary record found. Contact HR.")
        return redirect('myProfile')  # adjust to your actual self-profile url name

    buffer = _build_salary_slip_pdf(salary, profile, month, year)
    filename = f"salary_slip_{profile.user.username}_{year}_{month:02d}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def salary_slip_admin_view(request, uuid, year, month):
    if not user_has_permission(request.user, "manage_salary"):
        messages.error(
            request,
            "You don't have permission to do this."
        )
        return redirect("dashboard")

    profile = get_object_or_404(
        empProfile,
        uuid=uuid,
    )

    salary = profile.salary_records.filter(
        is_active=True
    ).first()

    if not salary:
        messages.error(
            request,
            "No salary record found for this employee."
        )
        return redirect(
            "editProfile",
            uuid=profile.uuid,
        )

    buffer = _build_salary_slip_pdf(
        salary,
        profile,
        month,
        year,
    )

    filename = (
        f"salary_slip_{profile.user.username}_{year}_{month:02d}.pdf"
    )

    response = HttpResponse(
        buffer,
        content_type="application/pdf",
    )

    response["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )

    return response

def _build_attendance_calendar(employee, year, month):
    month_calendar = calendar.Calendar(
        firstweekday=0
    )

    calendar_weeks = (
        month_calendar.monthdatescalendar(
            year,
            month,
        )
    )

    first_day = date(
        year,
        month,
        1,
    )

    if month == 12:
        next_month_first_day = date(
            year + 1,
            1,
            1,
        )
    else:
        next_month_first_day = date(
            year,
            month + 1,
            1,
        )

    # ==================================================
    # ATTENDANCE
    # ==================================================

    attendance_records = list(
        Attendance.objects.filter(
            employee=employee,
            date__gte=first_day,
            date__lt=next_month_first_day,
        )
    )

    attendance_by_date = {
        record.date: record
        for record in attendance_records
    }

    attendance_ids = [
        record.id
        for record in attendance_records
    ]

    # ==================================================
    # PENDING ATTENDANCE EXCEPTIONS
    # ==================================================

    pending_exceptions = (
        AttendanceException.objects
        .filter(
            attendance_id__in=attendance_ids,
            status="PENDING",
        )
    )

    pending_exception_by_attendance = {
        exception.attendance_id: exception
        for exception in pending_exceptions
    }

    # ==================================================
    # MANDATORY HOLIDAYS
    # ==================================================

    holidays = (
        Holiday.objects
        .filter(
            date__gte=first_day,
            date__lt=next_month_first_day,
            holiday_type="MANDATORY",
            is_active=True,
        )
        .order_by("date")
    )

    holiday_by_date = {
        holiday.date: holiday
        for holiday in holidays
    }

    # ==================================================
    # APPROVED OPTIONAL HOLIDAYS
    # Employee-specific
    # ==================================================

    approved_optional_requests = (
        OptionalHolidayRequest.objects
        .filter(
            employee=employee,
            status="APPROVED",
            holiday__holiday_type="OPTIONAL",
            holiday__is_active=True,
            holiday__date__gte=first_day,
            holiday__date__lt=next_month_first_day,
        )
        .select_related(
            "holiday"
        )
    )

    optional_holiday_by_date = {
        optional_request.holiday.date:
            optional_request
        for optional_request
        in approved_optional_requests
    }

    # ==================================================
    # APPROVED LEAVES
    # ==================================================

    approved_leaves = (
        LeaveRequest.objects
        .filter(
            employee=employee,
            status="APPROVED",
            start_date__lt=next_month_first_day,
            end_date__gte=first_day,
        )
        .select_related(
            "leave_type"
        )
        .order_by(
            "start_date"
        )
    )

    leave_by_date = {}

    for leave in approved_leaves:

        current_date = max(
            leave.start_date,
            first_day,
        )

        last_date = min(
            leave.end_date,
            next_month_first_day
            - timedelta(days=1),
        )

        while current_date <= last_date:

            leave_by_date[
                current_date
            ] = leave

            current_date += timedelta(
                days=1
            )

    # ==================================================
    # WORK SCHEDULE
    # ==================================================

    work_schedule = employee.work_schedule

    if not work_schedule:
        work_schedule = (
            WorkSchedule.objects
            .filter(
                is_default=True,
                is_active=True,
            )
            .first()
        )

    # ==================================================
    # ATTENDANCE STATUS CSS
    # ==================================================

    attendance_status_classes = {
        "PRESENT": "attendance-present",
        "LATE": "attendance-late",
        "HALF_DAY": "attendance-half-day",
        "ABSENT": "attendance-absent",
        "LEAVE": "attendance-leave",
    }

    weeks = []

    # ==================================================
    # BUILD DAYS
    # ==================================================

    for week in calendar_weeks:

        week_data = []

        for day in week:

            attendance = (
                attendance_by_date.get(day)
            )

            holiday = (
                holiday_by_date.get(day)
            )

            approved_optional_holiday = (
                optional_holiday_by_date.get(
                    day
                )
            )

            approved_leave = (
                leave_by_date.get(day)
            )

            pending_exception = None

            if attendance:
                pending_exception = (
                    pending_exception_by_attendance
                    .get(
                        attendance.id
                    )
                )

            # ==========================================
            # WEEKLY OFF
            # ==========================================

            is_weekly_off = False

            if work_schedule:
                is_weekly_off = not (
                    work_schedule
                    .is_working_day(day)
                )

            # ==========================================
            # DISPLAY PRIORITY
            #
            # 1 Mandatory Holiday
            # 2 Weekly Off
            # 3 Optional Holiday
            # 4 Approved Leave
            # 5 Attendance
            # ==========================================

            display_status = (
                "Not Marked"
            )

            status_class = ""

            filter_key = (
                "NOT_MARKED"
            )

            # ------------------------------------------
            # MANDATORY HOLIDAY
            # ------------------------------------------

            if holiday:

                display_status = "Holiday"

                status_class = (
                    "attendance-holiday"
                )

                filter_key = "HOLIDAY"

            # ------------------------------------------
            # WEEKLY OFF
            # ------------------------------------------

            elif is_weekly_off:

                display_status = (
                    "Weekly Off"
                )

                status_class = (
                    "attendance-weekly-off"
                )

                filter_key = (
                    "WEEKLY_OFF"
                )

            # ------------------------------------------
            # APPROVED OPTIONAL HOLIDAY
            # ------------------------------------------

            elif approved_optional_holiday:

                display_status = (
                    "Optional Holiday"
                )

                status_class = (
                    "attendance-optional-holiday"
                )

                filter_key = (
                    "OPTIONAL_HOLIDAY"
                )

            # ------------------------------------------
            # APPROVED LEAVE
            # ------------------------------------------

            elif approved_leave:

                display_status = "Leave"

                status_class = (
                    "attendance-leave"
                )

                filter_key = "LEAVE"

            # ------------------------------------------
            # ATTENDANCE
            # ------------------------------------------

            elif attendance:

                display_status = (
                    attendance
                    .get_status_display()
                )

                status_class = (
                    attendance_status_classes
                    .get(
                        attendance.status,
                        "",
                    )
                )

                filter_key = (
                    attendance.status
                )

                if pending_exception:
                    filter_key = (
                        "PENDING"
                    )

            # ==========================================
            # DAY DATA
            # ==========================================

            week_data.append({
                "date": day,

                "day_number": (
                    day.day
                ),

                "is_current_month": (
                    day.month == month
                ),

                "is_today": (
                    day
                    == timezone.localdate()
                ),

                "attendance": (
                    attendance
                ),

                "pending_exception": (
                    pending_exception
                ),

                "holiday": (
                    holiday
                ),

                "approved_optional_holiday": (
                    approved_optional_holiday
                ),

                "approved_leave": (
                    approved_leave
                ),

                "is_weekly_off": (
                    is_weekly_off
                ),

                "display_status": (
                    display_status
                ),

                "status_class": (
                    status_class
                ),

                "filter_key": (
                    filter_key
                ),
            })

        weeks.append(
            week_data
        )

    # ==================================================
    # PREVIOUS MONTH
    # ==================================================

    if month == 1:

        previous_year = (
            year - 1
        )

        previous_month = 12

    else:

        previous_year = year

        previous_month = (
            month - 1
        )

    # ==================================================
    # NEXT MONTH
    # ==================================================

    if month == 12:

        next_year = (
            year + 1
        )

        next_month = 1

    else:

        next_year = year

        next_month = (
            month + 1
        )

    return {
        "weeks": weeks,

        "month_name": (
            calendar.month_name[
                month
            ]
        ),

        "year": year,

        "month": month,

        "previous_year": (
            previous_year
        ),

        "previous_month": (
            previous_month
        ),

        "next_year": (
            next_year
        ),

        "next_month": (
            next_month
        ),
    }


@login_required
def attendance_calendar_self_view(request):
    """
    Employee views their own attendance calendar.
    """

    employee = get_object_or_404(
        empProfile.objects.select_related(
            "user",
            "manager__user",
            "designation",
            "shift",
            "work_schedule",
        ),
        user=request.user,
    )

    today = timezone.localdate()

    # ==================================================
    # MONTH / YEAR
    # ==================================================

    try:

        year = int(
            request.GET.get(
                "year",
                today.year,
            )
        )

        month = int(
            request.GET.get(
                "month",
                today.month,
            )
        )

        selected_date = date(
            year,
            month,
            1,
        )

    except (
        TypeError,
        ValueError,
    ):

        selected_date = date(
            today.year,
            today.month,
            1,
        )

    # ==================================================
    # VIEW MODE
    # calendar / list
    # ==================================================

    view_mode = request.GET.get(
        "view",
        "calendar",
    )

    if view_mode not in [
        "calendar",
        "list",
    ]:
        view_mode = "calendar"

    # ==================================================
    # FILTER
    # ==================================================

    status_filter = request.GET.get(
        "filter",
        "ALL",
    ).upper()

    allowed_filters = [
        "ALL",
        "PRESENT",
        "LATE",
        "HALF_DAY",
        "ABSENT",
        "LEAVE",
        "HOLIDAY",
        "OPTIONAL_HOLIDAY",
        "WEEKLY_OFF",
        "PENDING",
    ]

    if status_filter not in allowed_filters:
        status_filter = "ALL"

    # ==================================================
    # CALENDAR DATA
    # ==================================================

    calendar_data = (
        _build_attendance_calendar(
            employee,
            selected_date.year,
            selected_date.month,
        )
    )

    context = {
        "employee": employee,

        "is_admin_calendar": False,

        "view_mode": view_mode,

        "status_filter": (
            status_filter
        ),

        **calendar_data,
    }

    return render(
        request,
        "appEmp/attendance_calendar.html",
        context,
    )

@login_required
def raise_attendance_exception_view(request, attendance_uuid):
    employee = get_object_or_404(
        empProfile,
        user=request.user,
    )

    attendance = get_object_or_404(
        Attendance.objects.select_related("employee__user"),
        uuid=attendance_uuid,
        employee=employee,
    )

    calendar_url = reverse("attendanceCalendarSelf")
    calendar_redirect_url = (
        f"{calendar_url}"
        f"?year={attendance.date.year}"
        f"&month={attendance.date.month}"
    )

    if attendance.status != "ABSENT":
        messages.error(
            request,
            "An exception request can currently be raised only for an absent day."
        )
        return redirect(calendar_redirect_url)

    existing_pending_request = AttendanceException.objects.filter(
        attendance=attendance,
        raised_by=request.user,
        status="PENDING",
    ).exists()

    if existing_pending_request:
        messages.info(
            request,
            "A pending exception request already exists for this date."
        )
        return redirect(calendar_redirect_url)

    if request.method == "POST":
        form = AttendanceExceptionForm(request.POST)

        if form.is_valid():
            exception = form.save(commit=False)
            exception.attendance = attendance
            exception.raised_by = request.user
            exception.status = "PENDING"
            exception.save()

            messages.success(
                request,
                "Attendance exception request submitted successfully."
            )
            return redirect(calendar_redirect_url)

    else:
        form = AttendanceExceptionForm()

    return render(
        request,
        "appEmp/raise_attendance_exception.html",
        {
            "form": form,
            "attendance": attendance,
            "employee": employee,
        },
    )


@login_required
def attendance_calendar_admin_view(
    request,
    uuid,
):
    """
    HR/Admin or the employee's direct manager
    can view an employee's attendance calendar.
    """

    employee = get_object_or_404(
        empProfile.objects.select_related(
            "user",
            "manager__user",
            "designation",
            "shift",
            "work_schedule",
        ),
        uuid=uuid,
    )

    # ==================================================
    # CURRENT LOGGED-IN EMPLOYEE
    # ==================================================

    try:
        logged_in_profile = (
            empProfile.objects.get(
                user=request.user
            )
        )

    except empProfile.DoesNotExist:

        # Superusers may not necessarily
        # have an employee profile.
        logged_in_profile = None

    # ==================================================
    # PERMISSION
    # ==================================================

    is_admin = user_has_permission(
        request.user,
        "manage_attendance",
    )

    is_direct_manager = False

    if logged_in_profile:

        is_direct_manager = (
            employee.manager_id
            == logged_in_profile.id
        )

    if (
        not is_admin
        and not is_direct_manager
    ):

        messages.error(
            request,
            (
                "You don't have permission to "
                "view this attendance calendar."
            ),
        )

        return redirect(
            "dashboard"
        )

    # ==================================================
    # MONTH / YEAR
    # ==================================================

    today = timezone.localdate()

    try:

        year = int(
            request.GET.get(
                "year",
                today.year,
            )
        )

        month = int(
            request.GET.get(
                "month",
                today.month,
            )
        )

        selected_date = date(
            year,
            month,
            1,
        )

    except (
        TypeError,
        ValueError,
    ):

        selected_date = date(
            today.year,
            today.month,
            1,
        )

    # ==================================================
    # VIEW MODE
    # ==================================================

    view_mode = request.GET.get(
        "view",
        "calendar",
    )

    if view_mode not in [
        "calendar",
        "list",
    ]:
        view_mode = "calendar"

    # ==================================================
    # STATUS FILTER
    # ==================================================

    status_filter = request.GET.get(
        "filter",
        "ALL",
    ).upper()

    allowed_filters = [
        "ALL",
        "PRESENT",
        "LATE",
        "HALF_DAY",
        "ABSENT",
        "LEAVE",
        "HOLIDAY",
        "OPTIONAL_HOLIDAY",
        "WEEKLY_OFF",
        "PENDING",
    ]

    if status_filter not in allowed_filters:
        status_filter = "ALL"

    # ==================================================
    # CALENDAR DATA
    # ==================================================

    calendar_data = (
        _build_attendance_calendar(
            employee,
            selected_date.year,
            selected_date.month,
        )
    )

    context = {
        "employee": employee,

        "is_admin_calendar": True,

        "view_mode": view_mode,

        "status_filter": (
            status_filter
        ),

        **calendar_data,
    }

    return render(
        request,
        "appEmp/attendance_calendar.html",
        context,
    )


def _build_holiday_calendar(year, month):
    month_calendar = calendar.Calendar(firstweekday=0)
    calendar_weeks = month_calendar.monthdatescalendar(year, month)

    first_day = date(year, month, 1)

    if month == 12:
        next_month_first_day = date(year + 1, 1, 1)
    else:
        next_month_first_day = date(year, month + 1, 1)

    holidays = Holiday.objects.filter(
        date__gte=first_day,
        date__lt=next_month_first_day,
        is_active=True,
    ).order_by("date")

    holidays_by_date = {
        holiday.date: holiday
        for holiday in holidays
    }

    weeks = []

    for week in calendar_weeks:
        week_data = []

        for day in week:
            week_data.append({
                "date": day,
                "day_number": day.day,
                "is_current_month": day.month == month,
                "is_today": day == timezone.localdate(),
                "holiday": holidays_by_date.get(day),
            })

        weeks.append(week_data)

    if month == 1:
        previous_year = year - 1
        previous_month = 12
    else:
        previous_year = year
        previous_month = month - 1

    if month == 12:
        next_year = year + 1
        next_month = 1
    else:
        next_year = year
        next_month = month + 1

    return {
        "weeks": weeks,
        "month_name": calendar.month_name[month],
        "year": year,
        "month": month,
        "previous_year": previous_year,
        "previous_month": previous_month,
        "next_year": next_year,
        "next_month": next_month,
    }

def _build_holiday_year(year):
    month_calendar = calendar.Calendar(firstweekday=0)

    holidays = Holiday.objects.filter(
        date__year=year,
        is_active=True,
    ).order_by("date")

    holidays_by_date = {}

    for holiday in holidays:
        holidays_by_date.setdefault(
            holiday.date,
            []
        ).append(holiday)

    months = []

    for month in range(1, 13):
        calendar_weeks = month_calendar.monthdatescalendar(
            year,
            month,
        )

        weeks = []

        for week in calendar_weeks:
            week_data = []

            for day in week:
                week_data.append({
                    "date": day,
                    "day_number": day.day,
                    "is_current_month": day.month == month,
                    "is_today": day == timezone.localdate(),
                    "holidays": holidays_by_date.get(day, []),
                })

            weeks.append(week_data)

        months.append({
            "number": month,
            "name": calendar.month_name[month],
            "weeks": weeks,
        })

    return months


@login_required
def holiday_calendar_view(request):
    today = timezone.localdate()

    try:
        year = int(request.GET.get("year", today.year))
        month = int(request.GET.get("month", today.month))

        selected_date = date(
            year,
            month,
            1,
        )

    except (TypeError, ValueError):
        selected_date = date(
            today.year,
            today.month,
            1,
        )

    calendar_data = _build_holiday_calendar(
        selected_date.year,
        selected_date.month,
    )

    return render(
        request,
        "appEmp/holiday_calendar.html",
        calendar_data,
    )



@login_required
def apply_leave_view(request):
    employee = get_object_or_404(
        empProfile.objects.select_related(
            "user",
            "manager__user"
        ),
        user=request.user,
    )

    if request.method == "POST":
        form = LeaveRequestForm(
            request.POST,
            employee=employee,
        )

        if form.is_valid():
            leave_request = form.save(commit=False)
            leave_request.employee = employee
            leave_request.status = "PENDING"
            leave_request.save()

            messages.success(
                request,
                "Leave request submitted successfully."
            )

            return redirect("myLeaveRequests")


    else:
        form = LeaveRequestForm(
            employee=employee
        )

    current_year = timezone.localdate().year

    balances = (
        LeaveBalance.objects
        .filter(
            employee=employee,
            year=current_year
        )
        .select_related("leave_type")
        .order_by("leave_type__name")
    )

    return render(
        request,
        "appEmp/apply_leave.html",
        {
            "form": form,
            "employee": employee,
            "balances": balances,
            "current_year": current_year,
        },
    )


@login_required
def my_leave_requests_view(request):
    employee = get_object_or_404(
        empProfile,
        user=request.user,
    )

    leave_requests = (
        LeaveRequest.objects
        .filter(employee=employee)
        .select_related(
            "leave_type",
            "reviewed_by",
        )
        .order_by("-created_at")
    )

    current_year = timezone.localdate().year

    balances = (
        LeaveBalance.objects
        .filter(employee=employee, year=current_year)
        .select_related("leave_type")
        .order_by("leave_type__name")
    )

    return render(
        request,
        "appEmp/my_leave_requests.html",
        {
            "employee": employee,
            "leave_requests": leave_requests,
            "balances": balances,
            "current_year": current_year,
        },
    )


@login_required
def pending_leave_requests_view(request):
    logged_in_profile = get_object_or_404(
        empProfile,
        user=request.user,
    )

    if can_review_leave(request.user):
        leave_requests = LeaveRequest.objects.filter(
            status="PENDING"
        )
    else:
        leave_requests = LeaveRequest.objects.filter(
            status="PENDING",
            employee__manager=logged_in_profile,
        )

    leave_requests = (
        leave_requests
        .select_related(
            "employee__user",
            "employee__manager__user",
            "leave_type",
        )
        .order_by(
            "start_date",
            "created_at",
        )
    )

    return render(
        request,
        "appEmp/pending_leave_requests.html",
        {
            "leave_requests": leave_requests,
        },
    )


@login_required
def review_leave_request_view(request, uuid):
    reviewer_profile = get_object_or_404(
        empProfile,
        user=request.user,
    )

    leave_request = get_object_or_404(
        LeaveRequest.objects.select_related(
            "employee__user",
            "employee__manager__user",
            "leave_type",
        ),
        uuid=uuid,
    )

    is_admin = can_review_leave(
        request.user
    )

    is_direct_manager = (
        leave_request.employee.manager_id
        == reviewer_profile.id
    )

    if not is_admin and not is_direct_manager:
        messages.error(
            request,
            "You don't have permission to review this leave request."
        )
        return redirect("dashboard")

    if leave_request.status != "PENDING":
        messages.info(
            request,
            "This leave request has already been reviewed."
        )
        return redirect(
            "pendingLeaveRequests"
        )

    if request.method == "POST":
        form = LeaveRequestReviewForm(
            request.POST,
            instance=leave_request,
        )

        if form.is_valid():
            action = form.cleaned_data["action"]

            manager_comment = (
                form.cleaned_data.get(
                    "manager_comment"
                )
            )

            try:
                with transaction.atomic():

                    locked_request = (
                        LeaveRequest.objects
                        .select_for_update()
                        .select_related(
                            "employee",
                            "leave_type",
                        )
                        .get(
                            pk=leave_request.pk
                        )
                    )

                    if (
                        locked_request.status
                        != "PENDING"
                    ):
                        messages.info(
                            request,
                            "This request was already reviewed."
                        )

                        return redirect(
                            "pendingLeaveRequests"
                        )

                    if action == "APPROVE":

                        leave_balance = (
                            LeaveBalance.objects
                            .select_for_update()
                            .filter(
                                employee=locked_request.employee,
                                leave_type=locked_request.leave_type,
                                year=locked_request.start_date.year,
                            )
                            .first()
                        )

                        if not leave_balance:
                            messages.error(
                                request,
                                (
                                    "No leave balance exists for this "
                                    "employee, leave type and year."
                                )
                            )

                            return redirect(
                                "reviewLeaveRequest",
                                uuid=locked_request.uuid,
                            )

                        if (
                            leave_balance.available_balance
                            < locked_request.number_of_days
                        ):
                            messages.error(
                                request,
                                (
                                    "Insufficient leave balance. "
                                    f"Available: "
                                    f"{leave_balance.available_balance}, "
                                    f"Requested: "
                                    f"{locked_request.number_of_days}."
                                )
                            )

                            return redirect(
                                "reviewLeaveRequest",
                                uuid=locked_request.uuid,
                            )

                        leave_dates = []

                        current_date = (
                            locked_request.start_date
                        )

                        while (
                            current_date
                            <= locked_request.end_date
                        ):
                            leave_dates.append(
                                current_date
                            )

                            current_date += timedelta(
                                days=1
                            )

                        existing_attendance = {
                            attendance.date: attendance
                            for attendance in (
                                Attendance.objects
                                .select_for_update()
                                .filter(
                                    employee=locked_request.employee,
                                    date__in=leave_dates,
                                )
                            )
                        }

                        conflicting_dates = []

                        for leave_date in leave_dates:
                            attendance = (
                                existing_attendance.get(
                                    leave_date
                                )
                            )

                            if not attendance:
                                continue

                            has_work_record = (
                                attendance.check_in
                                is not None
                                or attendance.check_out
                                is not None
                                or attendance.status
                                in [
                                    "PRESENT",
                                    "LATE",
                                ]
                            )

                            if has_work_record:
                                conflicting_dates.append(
                                    leave_date
                                )

                        if conflicting_dates:

                            formatted_dates = ", ".join(
                                leave_date.strftime(
                                    "%d %b %Y"
                                )
                                for leave_date
                                in conflicting_dates
                            )

                            messages.error(
                                request,
                                (
                                    "Leave cannot be approved because "
                                    "attendance is already marked on: "
                                    f"{formatted_dates}."
                                )
                            )

                            return redirect(
                                "reviewLeaveRequest",
                                uuid=locked_request.uuid,
                            )

                        leave_balance.used = (
                            leave_balance.used
                            + locked_request.number_of_days
                        )

                        leave_balance.save(
                            update_fields=[
                                "used",
                                "updated_at",
                            ]
                        )

                        is_half_day = (
                            locked_request.duration_type
                            in [
                                "FIRST_HALF",
                                "SECOND_HALF",
                            ]
                        )

                        attendance_status = (
                            "HALF_DAY"
                            if is_half_day
                            else "LEAVE"
                        )

                        for leave_date in leave_dates:

                            attendance = (
                                existing_attendance.get(
                                    leave_date
                                )
                            )

                            if attendance:
                                attendance.status = (
                                    attendance_status
                                )

                                attendance.marked_by = (
                                    request.user
                                )

                                attendance.save(
                                    update_fields=[
                                        "status",
                                        "marked_by",
                                        "updated_at",
                                    ]
                                )

                            else:
                                Attendance.objects.create(
                                    employee=locked_request.employee,
                                    date=leave_date,
                                    status=attendance_status,
                                    marked_by=request.user,
                                )

                        locked_request.status = (
                            "APPROVED"
                        )

                        success_message = (
                            "Leave request approved successfully. "
                            "The leave balance and attendance calendar "
                            "have been updated."
                        )

                    else:
                        locked_request.status = (
                            "REJECTED"
                        )

                        success_message = (
                            "Leave request rejected successfully."
                        )

                    locked_request.reviewed_by = (
                        request.user
                    )

                    locked_request.reviewed_at = (
                        timezone.now()
                    )

                    locked_request.manager_comment = (
                        manager_comment
                    )

                    locked_request.save(
                        update_fields=[
                            "status",
                            "reviewed_by",
                            "reviewed_at",
                            "manager_comment",
                            "updated_at",
                        ]
                    )

                messages.success(
                    request,
                    success_message,
                )

                return redirect(
                    "pendingLeaveRequests"
                )

            except LeaveRequest.DoesNotExist:
                messages.error(
                    request,
                    "Leave request could not be found."
                )

                return redirect(
                    "pendingLeaveRequests"
                )

    else:
        form = LeaveRequestReviewForm(
            instance=leave_request
        )

    leave_balance = LeaveBalance.objects.filter(
        employee=leave_request.employee,
        leave_type=leave_request.leave_type,
        year=leave_request.start_date.year,
    ).first()

    return render(
        request,
        "appEmp/review_leave_request.html",
        {
            "leave_request": leave_request,
            "leave_balance": leave_balance,
            "form": form,
        },
    )


@login_required
def all_leave_requests_view(request):
    logged_in_profile = get_object_or_404(
        empProfile,
        user=request.user,
    )

    # HR/Admin only
    if not can_review_leave(request.user):
        messages.error(
            request,
            "You don't have permission to view all leave requests."
        )
        return redirect("dashboard")

    leave_requests = (
        LeaveRequest.objects
        .select_related(
            "employee__user",
            "employee__manager__user",
            "leave_type",
            "reviewed_by",
        )
        .all()
    )

    # ── Filters ─────────────────────────────────────────────

    employee_id = request.GET.get("employee")
    leave_type_id = request.GET.get("leave_type")
    status = request.GET.get("status")
    year = request.GET.get("year")

    if employee_id:
        leave_requests = leave_requests.filter(
            employee_id=employee_id
        )

    if leave_type_id:
        leave_requests = leave_requests.filter(
            leave_type_id=leave_type_id
        )

    if status:
        leave_requests = leave_requests.filter(
            status=status
        )

    if year:
        try:
            year = int(year)

            leave_requests = leave_requests.filter(
                start_date__year=year
            )
        except (TypeError, ValueError):
            year = None

    leave_requests = leave_requests.order_by(
        "-created_at"
    )

    employees = (
        empProfile.objects
        .filter(is_active=True)
        .select_related("user")
        .order_by(
            "user__first_name",
            "user__last_name"
        )
    )

    leave_types = (
        LeaveTypeMaster.objects
        .filter(is_active=True)
        .order_by("name")
    )

    current_year = timezone.localdate().year

    years = range(
        current_year + 1,
        current_year - 5,
        -1,
    )

    context = {
        "leave_requests": leave_requests,
        "employees": employees,
        "leave_types": leave_types,
        "years": years,

        # Keep selected filters after submit
        "selected_employee": employee_id,
        "selected_leave_type": leave_type_id,
        "selected_status": status,
        "selected_year": str(year) if year else "",
    }

    return render(
        request,
        "appEmp/all_leave_requests.html",
        context,
    )


@login_required
def leave_balance_list_view(request):
    logged_in_profile = get_object_or_404(
        empProfile,
        user=request.user,
    )

    if not user_has_permission(request.user, "manage_leave_balance"):
        messages.error(
            request,
            "You don't have permission to manage leave balances."
        )
        return redirect("dashboard")

    balances = (
        LeaveBalance.objects
        .select_related(
            "employee__user",
            "leave_type",
            "assigned_by",
        )
        .all()
    )

    employee_id = request.GET.get("employee")
    leave_type_id = request.GET.get("leave_type")
    year = request.GET.get("year")

    if employee_id:
        balances = balances.filter(employee_id=employee_id)

    if leave_type_id:
        balances = balances.filter(leave_type_id=leave_type_id)

    if year:
        try:
            balances = balances.filter(year=int(year))
        except (TypeError, ValueError):
            pass

    balances = balances.order_by(
        "-year",
        "employee__user__first_name",
        "leave_type__name",
    )

    employees = (
        empProfile.objects
        .filter(is_active=True)
        .select_related("user")
        .order_by("user__first_name", "user__last_name")
    )

    leave_types = (
        LeaveTypeMaster.objects
        .filter(is_active=True)
        .order_by("name")
    )

    current_year = timezone.localdate().year

    return render(
        request,
        "appEmp/leave_balance_list.html",
        {
            "balances": balances,
            "employees": employees,
            "leave_types": leave_types,
            "years": range(current_year + 1, current_year - 5, -1),
            "selected_employee": employee_id,
            "selected_leave_type": leave_type_id,
            "selected_year": year or "",
        },
    )


@login_required
def leave_balance_edit_view(request, uuid=None):

    if not user_has_permission(
        request.user,
        "manage_leave_balance"
    ):
        messages.error(
            request,
            (
                "You don't have permission "
                "to manage leave balances."
            ),
        )
        return redirect("dashboard")

    # -----------------------------------------------------
    # Editing existing leave balances is disabled.
    # -----------------------------------------------------

    if uuid:
        messages.error(
            request,
            (
                "Editing an assigned leave balance "
                "is not allowed."
            ),
        )
        return redirect("leaveBalanceList")

    if request.method == "POST":

        form = LeaveBalanceForm(
            request.POST
        )

        if form.is_valid():

            with transaction.atomic():

                record = form.save(
                    commit=False
                )

                record.assigned_by = (
                    request.user
                )

                employee = record.employee

                leave_start_mode = (
                    form.cleaned_data[
                        "leave_start_mode"
                    ]
                )

                selected_date = (
                    form.cleaned_data.get(
                        "selected_leave_start_date"
                    )
                )

                # -----------------------------------------
                # Leave facility eligibility
                # -----------------------------------------

                if leave_start_mode == "TODAY":

                    employee.leave_eligible_from = (
                        timezone.localdate()
                    )

                elif (
                    leave_start_mode
                    == "AFTER_PROBATION"
                ):

                    # NULL means:
                    # use date_hired + 3 calendar months.
                    employee.leave_eligible_from = None

                elif (
                    leave_start_mode
                    == "SELECTED_DATE"
                ):

                    employee.leave_eligible_from = (
                        selected_date
                    )

                employee.save(
                    update_fields=[
                        "leave_eligible_from",
                        "updated_at",
                    ]
                )

                record.save()

            eligible_from = (
                employee
                .effective_leave_eligible_from
            )

            messages.success(
                request,
                (
                    "Leave balance assigned successfully. "
                    "Leave facility will be available from "
                    f"{eligible_from.strftime('%d %b %Y')}."
                ),
            )

            return redirect(
                "leaveBalanceList"
            )

    else:
        form = LeaveBalanceForm()

    return render(
        request,
        "appEmp/leave_balance_form.html",
        {
            "form": form,
        },
    )


@login_required
def pending_attendance_exceptions_view(request):
    reviewer_profile = get_object_or_404(
        empProfile,
        user=request.user,
    )

    if can_review_attendance_exception(
        request.user
    ):
        exception_requests = (
            AttendanceException.objects
            .filter(
                status="PENDING"
            )
        )

    else:
        exception_requests = (
            AttendanceException.objects
            .filter(
                status="PENDING",
                attendance__employee__manager=(
                    reviewer_profile
                ),
            )
        )

    exception_requests = (
        exception_requests
        .select_related(
            "attendance",
            "attendance__employee__user",
            "attendance__employee__manager__user",
            "raised_by",
        )
        .order_by(
            "attendance__date",
            "created_at",
        )
    )

    return render(
        request,
        "appEmp/pending_attendance_exceptions.html",
        {
            "exception_requests": (
                exception_requests
            ),
        },
    )


@login_required
def review_attendance_exception_view(
    request,
    uuid,
):
    reviewer_profile = get_object_or_404(
        empProfile,
        user=request.user,
    )

    exception_request = get_object_or_404(
        AttendanceException.objects
        .select_related(
            "attendance",
            "attendance__employee__user",
            "attendance__employee__manager__user",
            "raised_by",
        ),
        uuid=uuid,
    )

    employee = (
        exception_request.attendance.employee
    )

    is_admin = can_review_attendance_exception(
        request.user
    )

    is_direct_manager = (
        employee.manager_id
        == reviewer_profile.id
    )

    if not is_admin and not is_direct_manager:
        messages.error(
            request,
            (
                "You don't have permission to review "
                "this attendance exception."
            )
        )

        return redirect(
            "dashboard"
        )

    if exception_request.status != "PENDING":
        messages.info(
            request,
            (
                "This attendance exception "
                "has already been reviewed."
            )
        )

        return redirect(
            "pendingAttendanceExceptions"
        )

    if request.method == "POST":

        form = AttendanceExceptionReviewForm(
            request.POST
        )

        if form.is_valid():

            action = (
                form.cleaned_data["action"]
            )

            manager_comment = (
                form.cleaned_data.get(
                    "manager_comment"
                )
            )

            try:
                with transaction.atomic():

                    locked_exception = (
                        AttendanceException.objects
                        .select_for_update()
                        .select_related(
                            "attendance",
                            "attendance__employee",
                        )
                        .get(
                            pk=exception_request.pk
                        )
                    )

                    if (
                        locked_exception.status
                        != "PENDING"
                    ):
                        messages.info(
                            request,
                            (
                                "This attendance exception "
                                "was already reviewed."
                            )
                        )

                        return redirect(
                            "pendingAttendanceExceptions"
                        )

                    attendance = (
                        Attendance.objects
                        .select_for_update()
                        .get(
                            pk=(
                                locked_exception
                                .attendance_id
                            )
                        )
                    )

                    if action == "APPROVE":

                        if (
                            attendance.status
                            != "ABSENT"
                        ):
                            messages.error(
                                request,
                                (
                                    "This attendance record is no longer "
                                    "marked Absent, so the exception "
                                    "cannot be approved automatically."
                                )
                            )

                            return redirect(
                                "reviewAttendanceException",
                                uuid=locked_exception.uuid,
                            )

                        attendance.status = (
                            "PRESENT"
                        )

                        attendance.marked_by = (
                            request.user
                        )

                        attendance.save(
                            update_fields=[
                                "status",
                                "marked_by",
                                "updated_at",
                            ]
                        )

                        locked_exception.status = (
                            "APPROVED"
                        )

                        success_message = (
                            "Attendance exception approved successfully. "
                            "Attendance has been marked Present."
                        )

                    else:
                        locked_exception.status = (
                            "REJECTED"
                        )

                        success_message = (
                            "Attendance exception rejected successfully."
                        )

                    locked_exception.reviewed_by = (
                        request.user
                    )

                    locked_exception.reviewed_at = (
                        timezone.now()
                    )

                    locked_exception.manager_comment = (
                        manager_comment
                    )

                    locked_exception.save(
                        update_fields=[
                            "status",
                            "reviewed_by",
                            "reviewed_at",
                            "manager_comment",
                            "updated_at",
                        ]
                    )

                messages.success(
                    request,
                    success_message,
                )

                return redirect(
                    "pendingAttendanceExceptions"
                )

            except AttendanceException.DoesNotExist:
                messages.error(
                    request,
                    (
                        "Attendance exception "
                        "could not be found."
                    )
                )

                return redirect(
                    "pendingAttendanceExceptions"
                )

    else:
        form = (
            AttendanceExceptionReviewForm()
        )

    return render(
        request,
        "appEmp/review_attendance_exception.html",
        {
            "exception_request": (
                exception_request
            ),

            "attendance": (
                exception_request.attendance
            ),

            "form": form,
        },
    )


@login_required
def submit_resignation_view(request):
    employee = get_object_or_404(
        empProfile.objects.select_related(
            "user",
            "manager__user",
        ),
        user=request.user,
    )

    if request.method == "POST":
        form = ResignationForm(
            request.POST
        )

        if form.is_valid():
            try:
                exit_request = submit_resignation(
                    employee=employee,
                    initiated_by=request.user,
                    reason=form.cleaned_data["reason"],
                    proposed_last_working_date=(
                        form.cleaned_data[
                            "proposed_last_working_date"
                        ]
                    ),
                )

            except ValidationError as exc:
                for error in exc.messages:
                    messages.error(
                        request,
                        error
                    )

            else:
                messages.success(
                    request,
                    (
                        "Resignation submitted successfully. "
                        f"Request ID: {exit_request.uuid}"
                    )
                )

                return redirect(
                    "myResignations"
                )

    else:
        form = ResignationForm()

    return render(
        request,
        "appEmp/submit_resignation.html",
        {
            "form": form,
            "employee": employee,
        },
    )

@login_required
def my_resignations_view(request):
    employee = get_object_or_404(
        empProfile,
        user=request.user,
    )

    exit_requests = (
        EmployeeExitRequest.objects
        .filter(
            employee=employee,
            exit_type="RESIGNATION",
        )
        .select_related(
            "initiated_by",
            "hr_reviewed_by",
            "clearance_request",
        )
        .prefetch_related(
            "clearance_request__approvals__assigned_to",
            "clearance_request__approvals__decided_by",
        )
        .order_by("-created_at")
    )

    request_rows = []

    for exit_request in exit_requests:

        manager_approval = None
        finance_approval = None

        if exit_request.clearance_request_id:
            approvals = {
                approval.approval_type: approval
                for approval
                in exit_request.clearance_request.approvals.all()
            }

            manager_approval = approvals.get(
                "MANAGER_NOC"
            )

            finance_approval = approvals.get(
                "FINANCE_NOC"
            )

        request_rows.append({
            "exit_request": exit_request,
            "manager_approval": manager_approval,
            "finance_approval": finance_approval,
        })

    return render(
        request,
        "appEmp/my_resignations.html",
        {
            "employee": employee,
            "request_rows": request_rows,
        },
    )


@login_required
@permission_required(
    "appEmp.view_all_clearance_requests",
    raise_exception=True,
)
def pending_resignations_view(request):
    pending_resignations = (
        EmployeeExitRequest.objects
        .filter(
            exit_type="RESIGNATION",
            status="SUBMITTED",
        )
        .select_related(
            "employee__user",
            "employee__manager__user",
            "initiated_by",
        )
        .order_by("created_at")
    )

    return render(
        request,
        "appEmp/pending_resignations.html",
        {
            "pending_resignations": pending_resignations,
        },
    )

@login_required
@permission_required(
    "appEmp.apply_clearance_for_employee",
    raise_exception=True,
)
def review_resignation_view(request, uuid):
    exit_request = get_object_or_404(
        EmployeeExitRequest.objects.select_related(
            "employee__user",
            "employee__manager__user",
            "initiated_by",
        ),
        uuid=uuid,
        exit_type="RESIGNATION",
    )

    if exit_request.status != "SUBMITTED":
        messages.error(
            request,
            "This resignation has already been reviewed.",
        )
        return redirect(
            "pendingResignations"
        )

    if request.method == "POST":
        form = ResignationReviewForm(
            request.POST
        )

        if form.is_valid():
            try:
                review_resignation(
                    exit_request=exit_request,
                    reviewed_by=request.user,
                    decision=form.cleaned_data["decision"],
                    remarks=form.cleaned_data["remarks"],
                    final_last_working_date=(
                        form.cleaned_data[
                            "final_last_working_date"
                        ]
                    ),
                )

            except ValidationError as exc:
                for error in exc.messages:
                    messages.error(
                        request,
                        error,
                    )

            else:
                if form.cleaned_data["decision"] == "APPROVED":
                    messages.success(
                        request,
                        (
                            "Resignation accepted successfully. "
                            "Manager and Finance clearance has started."
                        ),
                    )
                else:
                    messages.success(
                        request,
                        "Resignation rejected successfully.",
                    )

                return redirect(
                    "pendingResignations"
                )

    else:
        form = ResignationReviewForm(
            initial={
                "final_last_working_date": (
                    exit_request.proposed_last_working_date
                ),
            }
        )

    return render(
        request,
        "appEmp/review_resignation.html",
        {
            "exit_request": exit_request,
            "form": form,
        },
    )

@login_required
def manager_clearance_queue_view(request):
    pending_approvals = (
        ClearanceApproval.objects
        .filter(
            approval_type="MANAGER_NOC",
            status="PENDING",
            assigned_to=request.user,
            clearance_request__status="PENDING",
        )
        .select_related(
            "clearance_request",
            "clearance_request__employee__user",
            "clearance_request__employee__manager__user",
            "clearance_request__employee_exit_request",
        )
        .order_by("created_at")
    )

    return render(
        request,
        "appEmp/manager_clearance_queue.html",
        {
            "pending_approvals": pending_approvals,
        },
    )


@login_required
def manager_clearance_review_view(request, uuid):
    approval = get_object_or_404(
        ClearanceApproval.objects.select_related(
            "clearance_request",
            "clearance_request__employee__user",
            "clearance_request__employee__manager__user",
            "clearance_request__employee_exit_request",
        ),
        uuid=uuid,
        approval_type="MANAGER_NOC",
    )

    # Object-level authorization:
    # manager can only review approvals assigned to them.
    if approval.assigned_to_id != request.user.id:
        raise PermissionDenied(
            "You are not authorized to review this clearance."
        )

    if approval.status != "PENDING":
        messages.error(
            request,
            "This Manager NOC has already been reviewed.",
        )
        return redirect(
            "managerClearanceQueue"
        )

    if request.method == "POST":
        form = ClearanceApprovalDecisionForm(
            request.POST
        )

        if form.is_valid():
            try:
                decide_clearance_approval(
                    approval=approval,
                    decided_by=request.user,
                    decision=form.cleaned_data["decision"],
                    remarks=form.cleaned_data["remarks"],
                )

            except ValidationError as exc:
                for error in exc.messages:
                    messages.error(
                        request,
                        error,
                    )

            else:
                if form.cleaned_data["decision"] == "APPROVED":
                    messages.success(
                        request,
                        "Manager NOC approved successfully.",
                    )
                else:
                    messages.success(
                        request,
                        "Manager NOC rejected successfully.",
                    )

                return redirect(
                    "managerClearanceQueue"
                )

    else:
        form = ClearanceApprovalDecisionForm()

    return render(
        request,
        "appEmp/manager_clearance_review.html",
        {
            "approval": approval,
            "form": form,
        },
    )

@login_required
@permission_required(
    "appEmp.review_finance_noc",
    raise_exception=True,
)
def finance_clearance_queue_view(request):
    pending_approvals = (
        ClearanceApproval.objects
        .filter(
            approval_type="FINANCE_NOC",
            status="PENDING",
            clearance_request__status="PENDING",
        )
        .select_related(
            "clearance_request",
            "clearance_request__employee__user",
            "clearance_request__employee_exit_request",
        )
        .prefetch_related(
        "clearance_request__approvals"
        )
        .order_by("created_at")
    )

    return render(
        request,
        "appEmp/finance_clearance_queue.html",
        {
            "pending_approvals": pending_approvals,
        },
    )

@login_required
@permission_required(
    "appEmp.review_finance_noc",
    raise_exception=True,
)
def finance_clearance_review_view(request, uuid):
    approval = get_object_or_404(
        ClearanceApproval.objects.select_related(
            "clearance_request",
            "clearance_request__employee__user",
            "clearance_request__employee_exit_request",
        ),
        uuid=uuid,
        approval_type="FINANCE_NOC",
    )

    if approval.status != "PENDING":
        messages.error(
            request,
            "This Finance NOC has already been reviewed.",
        )
        return redirect(
            "financeClearanceQueue"
        )

    if request.method == "POST":
        form = ClearanceApprovalDecisionForm(
            request.POST
        )

        if form.is_valid():
            try:
                decide_clearance_approval(
                    approval=approval,
                    decided_by=request.user,
                    decision=form.cleaned_data["decision"],
                    remarks=form.cleaned_data["remarks"],
                )

            except ValidationError as exc:
                for error in exc.messages:
                    messages.error(
                        request,
                        error,
                    )

            else:
                if form.cleaned_data["decision"] == "APPROVED":
                    messages.success(
                        request,
                        "Finance NOC approved successfully.",
                    )
                else:
                    messages.success(
                        request,
                        "Finance NOC rejected successfully.",
                    )

                return redirect(
                    "financeClearanceQueue"
                )

    else:
        form = ClearanceApprovalDecisionForm()

    return render(
        request,
        "appEmp/finance_clearance_review.html",
        {
            "approval": approval,
            "form": form,
        },
    )

@login_required
@permission_required(
    "appEmp.finalize_employee_exit",
    raise_exception=True,
)
def ready_for_finalization_view(request):
    exit_requests = (
        EmployeeExitRequest.objects
        .filter(
            status="READY_FOR_FINALIZATION",
        )
        .select_related(
            "employee__user",
            "employee__manager__user",
            "clearance_request",
            "hr_reviewed_by",
        )
        .prefetch_related(
            "clearance_request__approvals",
        )
        .order_by(
            "final_last_working_date",
            "created_at",
        )
    )

    today = timezone.localdate()

    return render(
        request,
        "appEmp/ready_for_finalization.html",
        {
            "exit_requests": exit_requests,
            "today": today,
        },
    )

@login_required
@permission_required(
    "appEmp.finalize_employee_exit",
    raise_exception=True,
)
def finalize_employee_exit_view(request, uuid):
    exit_request = get_object_or_404(
        EmployeeExitRequest.objects.select_related(
            "employee__user",
            "clearance_request",
        ),
        uuid=uuid,
    )

    if request.method != "POST":
        raise PermissionDenied(
            "Employee exit finalization must be submitted securely."
        )

    try:
        finalize_employee_exit(
            exit_request=exit_request,
            finalized_by=request.user,
        )

    except ValidationError as exc:
        for error in exc.messages:
            messages.error(
                request,
                error,
            )

    else:
        messages.success(
            request,
            (
                f"Exit finalized successfully for "
                f"{exit_request.employee.user.get_full_name() or exit_request.employee.user.username}."
            ),
        )

    return redirect(
        "readyForFinalization"
    )

@login_required
@permission_required(
    "appEmp.initiate_employee_exit",
    raise_exception=True,
)
def initiate_termination_view(request):

    if request.method == "POST":
        form = TerminationForm(request.POST)

        if form.is_valid():
            try:
                exit_request = initiate_termination(
                    employee=form.cleaned_data["employee"],
                    initiated_by=request.user,
                    reason=form.cleaned_data["reason"],
                    final_last_working_date=(
                        form.cleaned_data[
                            "final_last_working_date"
                        ]
                    ),
                )

            except ValidationError as exc:
                for error in exc.messages:
                    messages.error(
                        request,
                        error,
                    )

            else:
                messages.success(
                    request,
                    (
                        "Employee termination initiated successfully. "
                        "Manager and Finance clearances have started."
                    ),
                )

                return redirect(
                    "readyForFinalization"
                )

    else:
        form = TerminationForm()

    return render(
        request,
        "appEmp/initiate_termination.html",
        {
            "form": form,
        },
    )


@login_required
@permission_required(
    "appEmp.view_all_clearance_requests",
    raise_exception=True,
)
def all_exit_requests_view(request):
    selected_status = request.GET.get(
        "status",
        ""
    )

    selected_type = request.GET.get(
        "type",
        ""
    )

    exit_requests = (
        EmployeeExitRequest.objects
        .select_related(
            "employee__user",
            "employee__manager__user",
            "initiated_by",
            "hr_reviewed_by",
            "finalized_by",
            "clearance_request",
        )
        .prefetch_related(
            "clearance_request__approvals__assigned_to",
            "clearance_request__approvals__decided_by",
        )
        .order_by("-created_at")
    )

    if selected_status:
        exit_requests = exit_requests.filter(
            status=selected_status
        )

    if selected_type:
        exit_requests = exit_requests.filter(
            exit_type=selected_type
        )

    request_rows = []

    for exit_request in exit_requests:
        manager_approval = None
        finance_approval = None

        if exit_request.clearance_request_id:
            approvals = {
                approval.approval_type: approval
                for approval
                in exit_request.clearance_request.approvals.all()
            }

            manager_approval = approvals.get(
                "MANAGER_NOC"
            )

            finance_approval = approvals.get(
                "FINANCE_NOC"
            )

        request_rows.append({
            "exit_request": exit_request,
            "manager_approval": manager_approval,
            "finance_approval": finance_approval,
        })

    return render(
        request,
        "appEmp/all_exit_requests.html",
        {
            "request_rows": request_rows,
            "selected_status": selected_status,
            "selected_type": selected_type,

            "status_choices": (
                EmployeeExitRequest._meta
                .get_field("status")
                .choices
            ),

            "type_choices": (
                EmployeeExitRequest._meta
                .get_field("exit_type")
                .choices
            ),
        },
    )


@login_required
def holiday_list_view(request):
    today = timezone.localdate()

    try:
        year = int(request.GET.get("year", today.year))
    except (TypeError, ValueError):
        year = today.year

    holiday_type = request.GET.get("type", "")

    holidays = Holiday.objects.filter(
        date__year=year
    ).order_by("date")

    if holiday_type in ["MANDATORY", "OPTIONAL"]:
        holidays = holidays.filter(
            holiday_type=holiday_type
        )

    context = {
        "holidays": holidays,
        "year": year,
        "holiday_type": holiday_type,
    }

    return render(
        request,
        "appEmp/holiday_list.html",
        context,
    )


@login_required
def holiday_create_view(request):
    if request.method == "POST":
        form = HolidayForm(request.POST)

        if form.is_valid():
            holiday = form.save(commit=False)
            holiday.created_by = request.user
            holiday.save()

            messages.success(
                request,
                "Holiday added successfully."
            )

            return redirect("holidayList")

    else:
        form = HolidayForm()

    return render(
        request,
        "appEmp/holiday_form.html",
        {
            "form": form,
            "page_title": "Add Holiday",
        },
    )


@login_required
def holiday_edit_view(request, uuid):
    holiday = get_object_or_404(
        Holiday,
        uuid=uuid,
    )

    if request.method == "POST":
        form = HolidayForm(
            request.POST,
            instance=holiday,
        )

        if form.is_valid():
            form.save()

            messages.success(
                request,
                "Holiday updated successfully."
            )

            return redirect("holidayList")

    else:
        form = HolidayForm(
            instance=holiday
        )

    return render(
        request,
        "appEmp/holiday_form.html",
        {
            "form": form,
            "page_title": "Edit Holiday",
            "holiday": holiday,
        },
    )

@login_required
def holiday_year_view(request):
    today = timezone.localdate()

    try:
        year = int(
            request.GET.get("year", today.year)
        )
    except (TypeError, ValueError):
        year = today.year

    months = _build_holiday_year(year)

    default_month = (
        today.month
        if year == today.year
        else 1
    )

    return render(
        request,
        "appEmp/holiday_year.html",
        {
            "year": year,
            "months": months,
            "previous_year": year - 1,
            "next_year": year + 1,
            "default_month": default_month,
        },
    )

@login_required
def holiday_bulk_upload_view(request):
    preview_rows = []
    has_errors = False
    ready_count = 0
    skip_count = 0

    # --------------------------------------------------
    # FINAL IMPORT
    # --------------------------------------------------
    if (
        request.method == "POST"
        and request.POST.get("action") == "confirm_import"
    ):
        import_rows = request.session.get(
            "holiday_import_rows",
            []
        )

        if not import_rows:
            messages.error(
                request,
                "No validated holidays are available for import."
            )
            return redirect("holidayBulkUpload")

        imported_count = 0
        skipped_count = 0

        with transaction.atomic():

            for row in import_rows:
                name = row["name"]
                holiday_date = datetime.strptime(
                    row["date"],
                    "%Y-%m-%d",
                ).date()

                # Safety check in case record was added
                # after preview but before confirmation.
                existing = Holiday.objects.filter(
                    name__iexact=name,
                    date=holiday_date,
                ).exists()

                if existing:
                    skipped_count += 1
                    continue

                Holiday.objects.create(
                    name=name,
                    date=holiday_date,
                    holiday_type=row["holiday_type"],
                    description=row.get(
                        "description",
                        "",
                    ),
                    is_active=True,
                    created_by=request.user,
                )

                imported_count += 1

        request.session.pop(
            "holiday_import_rows",
            None,
        )

        messages.success(
            request,
            f"{imported_count} holiday(s) imported successfully. "
            f"{skipped_count} duplicate(s) skipped."
        )

        return redirect("holidayList")

    # --------------------------------------------------
    # FILE VALIDATION / PREVIEW
    # --------------------------------------------------
    if request.method == "POST":

        form = HolidayBulkUploadForm(
            request.POST,
            request.FILES,
        )

        if form.is_valid():

            uploaded_file = form.cleaned_data[
                "file"
            ]

            try:
                rows = []

                # ============================
                # XLSX
                # ============================
                if uploaded_file.name.lower().endswith(
                    ".xlsx"
                ):

                    workbook = load_workbook(
                        uploaded_file,
                        data_only=True,
                    )

                    worksheet = workbook.active

                    for row in worksheet.iter_rows(
                        values_only=True
                    ):
                        rows.append(list(row))

                # ============================
                # CSV
                # ============================
                elif uploaded_file.name.lower().endswith(
                    ".csv"
                ):

                    uploaded_file.seek(0)

                    decoded_file = (
                        uploaded_file
                        .read()
                        .decode("utf-8-sig")
                    )

                    reader = csv.reader(
                        io.StringIO(decoded_file)
                    )

                    rows = list(reader)

                # ============================
                # EMPTY FILE
                # ============================
                if not rows:

                    messages.error(
                        request,
                        "The uploaded file is empty."
                    )

                else:

                    expected_headers = [
                        "Holiday Name",
                        "Date",
                        "Type",
                        "Description",
                    ]

                    headers = [
                        str(value).strip()
                        if value is not None
                        else ""
                        for value in rows[0][:4]
                    ]

                    # ============================
                    # HEADER VALIDATION
                    # ============================
                    if headers != expected_headers:

                        messages.error(
                            request,
                            "Invalid file format. "
                            "Expected columns: Holiday Name, "
                            "Date, Type, Description."
                        )

                    else:

                        valid_import_rows = []

                        # Prevent duplicates inside
                        # the uploaded file itself.
                        seen_upload_rows = set()

                        for row_number, row in enumerate(
                            rows[1:],
                            start=2,
                        ):

                            # Ignore empty Excel rows
                            if not any(
                                value is not None
                                and str(value).strip()
                                for value in row
                            ):
                                continue

                            # -----------------------
                            # Values
                            # -----------------------
                            name = (
                                str(row[0]).strip()
                                if len(row) > 0
                                and row[0] is not None
                                else ""
                            )

                            raw_date = (
                                row[1]
                                if len(row) > 1
                                else None
                            )

                            holiday_type = (
                                str(row[2])
                                .strip()
                                .upper()
                                if len(row) > 2
                                and row[2] is not None
                                else ""
                            )

                            description = (
                                str(row[3]).strip()
                                if len(row) > 3
                                and row[3] is not None
                                else ""
                            )

                            errors = []
                            warnings = []

                            holiday_date = None

                            # ============================
                            # NAME
                            # ============================
                            if not name:
                                errors.append(
                                    "Holiday name is required."
                                )

                            # ============================
                            # DATE
                            # ============================
                            if isinstance(
                                raw_date,
                                datetime,
                            ):
                                holiday_date = (
                                    raw_date.date()
                                )

                            elif isinstance(
                                raw_date,
                                date,
                            ):
                                holiday_date = raw_date

                            elif raw_date:

                                raw_date_string = str(
                                    raw_date
                                ).strip()

                                accepted_formats = [
                                    "%Y-%m-%d",
                                    "%d-%m-%Y",
                                    "%d/%m/%Y",
                                ]

                                for date_format in (
                                    accepted_formats
                                ):
                                    try:

                                        holiday_date = (
                                            datetime.strptime(
                                                raw_date_string,
                                                date_format,
                                            ).date()
                                        )

                                        break

                                    except ValueError:
                                        continue

                            if not holiday_date:
                                errors.append(
                                    "Invalid date."
                                )

                            # ============================
                            # TYPE
                            # ============================
                            if holiday_type not in [
                                "MANDATORY",
                                "OPTIONAL",
                            ]:
                                errors.append(
                                    "Type must be MANDATORY "
                                    "or OPTIONAL."
                                )

                            # ============================
                            # DUPLICATE CHECK
                            # ============================
                            is_duplicate = False

                            if (
                                name
                                and holiday_date
                                and not errors
                            ):

                                duplicate_key = (
                                    name.lower(),
                                    holiday_date.isoformat(),
                                )

                                # Duplicate inside Excel
                                if duplicate_key in (
                                    seen_upload_rows
                                ):
                                    is_duplicate = True

                                    warnings.append(
                                        "Duplicate inside uploaded "
                                        "file — will be skipped."
                                    )

                                else:
                                    seen_upload_rows.add(
                                        duplicate_key
                                    )

                                    # Duplicate already DB
                                    if Holiday.objects.filter(
                                        name__iexact=name,
                                        date=holiday_date,
                                    ).exists():

                                        is_duplicate = True

                                        warnings.append(
                                            "Holiday already exists "
                                            "— will be skipped."
                                        )

                            # ============================
                            # STATUS
                            # ============================
                            if errors:
                                status = "ERROR"
                                has_errors = True

                            elif is_duplicate:
                                status = "SKIP"
                                skip_count += 1

                            else:
                                status = "READY"
                                ready_count += 1

                                valid_import_rows.append({
                                    "name": name,
                                    "date": (
                                        holiday_date
                                        .isoformat()
                                    ),
                                    "holiday_type": (
                                        holiday_type
                                    ),
                                    "description": (
                                        description
                                    ),
                                })

                            preview_rows.append({
                                "row_number": row_number,
                                "name": name,
                                "date": holiday_date,
                                "holiday_type": (
                                    holiday_type
                                ),
                                "description": description,
                                "errors": errors,
                                "warnings": warnings,
                                "status": status,
                            })

                        # Invalid row means we don't
                        # allow confirmation.
                        if has_errors:
                            request.session.pop(
                                "holiday_import_rows",
                                None,
                            )

                        else:
                            request.session[
                                "holiday_import_rows"
                            ] = valid_import_rows

            except Exception as exc:

                request.session.pop(
                    "holiday_import_rows",
                    None,
                )

                messages.error(
                    request,
                    f"Could not read the uploaded file: {exc}"
                )

    else:
        form = HolidayBulkUploadForm()

    return render(
        request,
        "appEmp/holiday_bulk_upload.html",
        {
            "form": form,
            "preview_rows": preview_rows,
            "has_errors": has_errors,
            "ready_count": ready_count,
            "skip_count": skip_count,
        },
    )

@login_required
def work_schedule_list_view(request):
    schedules = WorkSchedule.objects.all().order_by("name")

    return render(
        request,
        "appEmp/work_schedule_list.html",
        {
            "schedules": schedules,
        },
    )

@login_required
def work_schedule_create_view(request):
    if request.method == "POST":
        form = WorkScheduleForm(request.POST)

        if form.is_valid():
            form.save()

            messages.success(
                request,
                "Work schedule created successfully."
            )

            return redirect("workScheduleList")

    else:
        form = WorkScheduleForm()

    return render(
        request,
        "appEmp/work_schedule_form.html",
        {
            "form": form,
            "page_title": "Add Work Schedule",
        },
    )

@login_required
def work_schedule_edit_view(request, uuid):
    schedule = get_object_or_404(
        WorkSchedule,
        uuid=uuid,
    )

    if request.method == "POST":
        form = WorkScheduleForm(
            request.POST,
            instance=schedule,
        )

        if form.is_valid():
            form.save()

            messages.success(
                request,
                "Work schedule updated successfully."
            )

            return redirect("workScheduleList")

    else:
        form = WorkScheduleForm(
            instance=schedule
        )

    return render(
        request,
        "appEmp/work_schedule_form.html",
        {
            "form": form,
            "page_title": "Edit Work Schedule",
            "schedule": schedule,
        },
    )

@login_required
def optional_holiday_apply_view(request):
    employee = get_object_or_404(
        empProfile.objects.select_related(
            "user",
            "manager__user",
        ),
        user=request.user,
    )

    if request.method == "POST":
        form = OptionalHolidayRequestForm(
            request.POST,
            employee=employee,
        )

        if form.is_valid():
            optional_request = form.save(
                commit=False
            )

            optional_request.employee = employee
            optional_request.status = "PENDING"

            optional_request.save()

            messages.success(
                request,
                (
                    "Optional holiday request "
                    "submitted successfully."
                ),
            )

            return redirect(
                "optionalHolidayApply"
            )

    else:
        form = OptionalHolidayRequestForm(
            employee=employee
        )

    requests = (
        OptionalHolidayRequest.objects
        .filter(
            employee=employee
        )
        .select_related(
            "holiday",
            "reviewed_by",
        )
        .order_by(
            "-holiday__date"
        )
    )

    return render(
        request,
        "appEmp/optional_holiday_apply.html",
        {
            "form": form,
            "employee": employee,
            "optional_requests": requests,
        },
    )

@login_required
def pending_optional_holiday_requests_view(request):
    if not can_review_leave(request.user):
        messages.error(
            request,
            "You don't have permission to review optional holiday requests."
        )
        return redirect("dashboard")

    requests = (
        OptionalHolidayRequest.objects
        .filter(status="PENDING")
        .select_related(
            "employee__user",
            "employee__manager__user",
            "holiday",
        )
        .order_by("holiday__date")
    )

    return render(
        request,
        "appEmp/pending_optional_holiday_requests.html",
        {
            "optional_requests": requests,
        },
    )

@login_required
def review_optional_holiday_request_view(request, uuid):
    if not can_review_leave(request.user):
        messages.error(
            request,
            "You don't have permission to review this request."
        )
        return redirect("dashboard")

    optional_request = get_object_or_404(
        OptionalHolidayRequest.objects.select_related(
            "employee__user",
            "holiday",
        ),
        uuid=uuid,
    )

    if optional_request.status != "PENDING":
        messages.info(
            request,
            "This request has already been reviewed."
        )
        return redirect(
            "pendingOptionalHolidayRequests"
        )

    if request.method == "POST":
        action = request.POST.get("action")

        if action not in ["approve", "reject"]:
            messages.error(
                request,
                "Invalid review action."
            )
            return redirect(
                "pendingOptionalHolidayRequests"
            )

        optional_request.status = (
            "APPROVED"
            if action == "approve"
            else "REJECTED"
        )

        optional_request.reviewed_by = request.user
        optional_request.reviewed_at = timezone.now()

        optional_request.save(
            update_fields=[
                "status",
                "reviewed_by",
                "reviewed_at",
                "updated_at",
            ]
        )

        messages.success(
            request,
            (
                f"Optional holiday request "
                f"{optional_request.status.lower()} successfully."
            )
        )

    return redirect(
        "pendingOptionalHolidayRequests"
    )


@csrf_exempt
def whatsapp_webhook(request):
    if request.method == 'GET':
        mode = request.GET.get('hub.mode')
        token = request.GET.get('hub.verify_token')
        challenge = request.GET.get('hub.challenge')

        if mode == 'subscribe' and token == settings.WHATSAPP_VERIFY_TOKEN:
            return HttpResponse(challenge, status=200)
        return HttpResponse('Verification failed', status=403)

    if request.method == 'POST':
        data = json.loads(request.body)

        try:
            entry = data['entry'][0]
            changes = entry['changes'][0]['value']

            if 'messages' not in changes:
                return JsonResponse({'status': 'ignored'}, status=200)

            message = changes['messages'][0]
            contact = changes['contacts'][0]

            wa_message_id = message['id']
            sender_wa_id = message['from']
            sender_name = contact.get('profile', {}).get('name', '')
            message_body = message.get('text', {}).get('body', '')

            category = classify_message(message_body)

            ticket, created = WhatsAppTicket.objects.get_or_create(
                wa_message_id=wa_message_id,
                defaults={
                    'sender_wa_id': sender_wa_id,
                    'sender_name': sender_name,
                    'message_body': message_body,
                    'category': category,
                }
            )

            if created:
                reply_text = None

                if category == "leave":
                    employee = find_employee_by_phone(sender_wa_id)
                    if employee:
                        reply_text = get_leave_balance_reply(employee)
                        ticket.status = "auto_resolved"
                        ticket.save(update_fields=["status"])

                elif category == "attendance":
                    employee = find_employee_by_phone(sender_wa_id)
                    if employee:
                        reply_text = get_attendance_reply(employee)
                        ticket.status = "auto_resolved"
                        ticket.save(update_fields=["status"])

                elif category == "payroll":
                    employee = find_employee_by_phone(sender_wa_id)
                    if employee:
                        reply_text = get_payroll_reply(employee)
                        ticket.status = "auto_resolved"
                        ticket.save(update_fields=["status"])

                if not reply_text:
                    reply_text = f"Thanks {sender_name or ''}, we've logged your message as ticket #{ticket.id}. Our HR team will get back to you."

                send_whatsapp_message(sender_wa_id, reply_text)

        except (KeyError, IndexError):
            pass

        return JsonResponse({'status': 'ok'}, status=200)
    